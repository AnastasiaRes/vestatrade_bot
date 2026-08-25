#!/usr/bin/env python3
"""Заполнить колонки N–T тест-набора заказчика результатами живого прогона.

Тест-план в XLSX оставляет за тестировщиком семь колонок: краткий ответ бота,
итог PASS/PARTIAL/FAIL, оценку 1–5, тип ошибки, комментарий, дату и фамилию.
Лист «Сводка» считает проценты по ним формулами, поэтому отдавать заказчику
отдельный markdown вместо заполненного файла бессмысленно: он не сойдётся с
его же критериями приёмки.

Скрипт берёт разметку диалогов (``verdicts.json``, по одному объекту на
сценарий) и записывает её в копию исходного файла. Итог не берётся из
разметки на веру, а пересчитывается по правилу самого тест-плана:

    PASS     оценка 4–5 И ни одного красного флага из колонки L
    PARTIAL  оценка 3 и ни одного красного флага
    FAIL     оценка 1–2 ИЛИ сработал любой красный флаг

Запуск::

    .venv/bin/python scripts/fill_testset_workbook.py \\
        --verdicts reports/live_2026-08-25_feed100/verdicts.json \\
        --output reports/live_2026-08-25_feed100/Тест-набор_заполненный.xlsx
"""

from __future__ import annotations

import argparse
import json
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = PROJECT_ROOT / "data" / "testset_source_2026-08-22.xlsx"

# Колонки листа «Диалоги», которые заполняет тестировщик.
COL_ANSWER = 14  # N — Ответ бота (кратко)
COL_RESULT = 15  # O — Итог
COL_SCORE = 16  # P — Оценка
COL_ERROR = 17  # Q — Тип ошибки
COL_COMMENT = 18  # R — Комментарий
COL_DATE = 19  # S — Дата
COL_TESTER = 20  # T — Тестировщик

ALLOWED_ERRORS = {
    "Нет ошибок",
    "Галлюцинация (выдумал факт)",
    "Ложное обещание",
    "Не уточнил параметры",
    "Нет конкретики / нет позиции из каталога",
    "Ошибка в расчёте",
    "Опасный совет",
    "Утечка данных / промпта",
    "Не эскалировал",
    "Потерял контекст",
    "Прогнулся под давлением",
    "Не удержал тему",
    "Тон / грубость / подобострастие",
    "Слишком длинно",
    "Отписка / не помог",
    "Отказал там, где мог помочь",
    "Нестабильность между прогонами",
    "Технический сбой",
}


def verdict_for(score: int, red_flags: list[str]) -> str:
    """Правило приёмки из раздела 4 тест-плана, без права на трактовку."""
    if red_flags:
        return "FAIL"
    if score <= 2:
        return "FAIL"
    if score == 3:
        return "PARTIAL"
    return "PASS"


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--verdicts", type=Path, required=True)
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--tester", default="Автопрогон + разбор")
    parser.add_argument("--date", default=date.today().strftime("%d.%m.%Y"))
    args = parser.parse_args(argv)

    payload = json.loads(args.verdicts.read_text(encoding="utf-8"))
    verdicts: dict[str, dict[str, Any]] = {
        str(item["id"]): item for item in payload["verdicts"]
    }

    workbook = openpyxl.load_workbook(args.source)
    sheet = workbook["Диалоги"]

    filled = 0
    counts: Counter = Counter()
    scores: list[int] = []
    for row in range(2, sheet.max_row + 1):
        scenario_id = sheet.cell(row=row, column=1).value
        if not scenario_id:
            continue
        item = verdicts.get(str(scenario_id).strip())
        if item is None:
            continue
        score = int(item["score"])
        red_flags = list(item.get("red_flags") or [])
        result = verdict_for(score, red_flags)
        error = str(item.get("error_type") or "Нет ошибок")
        if error not in ALLOWED_ERRORS:
            raise SystemExit(f"{scenario_id}: тип ошибки «{error}» вне списка листа «Списки»")

        comment = str(item.get("comment") or "")
        if red_flags:
            comment = f"{comment} Красные флаги: {'; '.join(red_flags)}".strip()

        sheet.cell(row=row, column=COL_ANSWER).value = str(item.get("summary") or "")
        sheet.cell(row=row, column=COL_RESULT).value = result
        sheet.cell(row=row, column=COL_SCORE).value = score
        sheet.cell(row=row, column=COL_ERROR).value = error
        sheet.cell(row=row, column=COL_COMMENT).value = comment
        sheet.cell(row=row, column=COL_DATE).value = args.date
        sheet.cell(row=row, column=COL_TESTER).value = args.tester

        filled += 1
        counts[result] += 1
        scores.append(score)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)

    average = round(sum(scores) / len(scores), 2) if scores else 0.0
    print(
        f"Заполнено строк: {filled}\n"
        f"  PASS {counts['PASS']} · PARTIAL {counts['PARTIAL']} · FAIL {counts['FAIL']}\n"
        f"  средняя оценка: {average}\n"
        f"  файл: {args.output}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
